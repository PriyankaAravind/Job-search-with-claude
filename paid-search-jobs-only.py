"""
Daily Paid Search - Job Search
- Searches jobs via Adzuna API (free, 250 requests/day)
- Emails a formatted Excel report every morning
- No AI, no resume, no cost
"""

import os
import time
import smtplib
import requests
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from datetime import datetime, timezone, timedelta
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Config — edit these to match your profile ─────────────────────────────────

SKILLS = [ "GA4","looker studio", "GTM","microsoft ads", "full funnel", "campaign management", "client management", "bid management", 
        "budget management","google ads","google adwords", "bing ads","google ads editor", "google sheets","Budget", "bid",
        "presentation","insights","campaign optimization"]

SEARCH_QUERIES = ["Senior Paid Search Specialist","Senior SEM Specialist","Senior PPC Specialist","Search Marketing Analyst","Senior Paid Search Strategist",
    "Paid Search Manager","SEM","Paid Search", "Performance marketing","SEM Manager","Marketing analyst","Analytics & Insights", "search Campaign", 
                  "Paid media","PPC Manager","search engine marketing","search""growth marketing",
                  "Senior Paid Search Manager","SEM Account Manager","Performance Marketing Manager","paid media strategist"]

SKILLS = list(dict.fromkeys(SKILLS))
SEARCH_QUERIES = list(dict.fromkeys(SEARCH_QUERIES))

# ── Constants ──────────────────────────────────────────────────────────────────

OUTPUT_DIR       = Path("output")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
ADZUNA_BASE      = "https://api.adzuna.com/v1/api/jobs/ca/search"
HOURS_BACK       = 24
RESULTS_PER_PAGE = 20

# ── Job Fetching ───────────────────────────────────────────────────────────────
def fetch_jobs(app_id: str, app_key: str) -> list[dict]:
    all_jobs, seen_ids = [], set()
    cutoff = datetime.now(timezone.utc) - timedelta(hours=HOURS_BACK)

    for query in SEARCH_QUERIES:
        print(f"  Searching: '{query}' ...")
        for page in range(1, 4):
            params = {
                "app_id": app_id, "app_key": app_key, "what": query,"where": "toronto","distance":50,
                "results_per_page": RESULTS_PER_PAGE,
                "sort_by": "date", "max_days_old": 1,
            }
            try:
                resp = requests.get(f"{ADZUNA_BASE}/{page}",
                                    params=params, timeout=30)
                resp.raise_for_status()
                jobs = resp.json().get("results", [])
                if not jobs:
                    break
                added = 0
                for job in jobs:
                    jid = job.get("id")
                    if not jid or jid in seen_ids:
                        continue
                    try:
                        dt = datetime.fromisoformat(
                            job.get("created", "").replace("Z", "+00:00"))
                        if dt < cutoff:
                            continue
                    except Exception:
                        pass
                    seen_ids.add(jid)
                    all_jobs.append(job)
                    added += 1
                print(f"    Page {page}: {added} new jobs (total: {len(all_jobs)})")
                if len(jobs) < RESULTS_PER_PAGE:
                    break
                time.sleep(1)
            except Exception as e:
                print(f"    Warning: page {page} failed - {e}")
                time.sleep(3)
                break

    print(f"  Total unique jobs fetched: {len(all_jobs)}")
    return all_jobs


# ── Data Processing ────────────────────────────────────────────────────────────

remote_terms = ["remote","hybrid","work from home","wfh","flexible","remote first","fully remote"]

def is_remote_job(title, desc):
    text = f"{title} {desc}".lower()
    return any(term in text for term in remote_terms)
              
def score_job(job: dict) -> int:
    score=0
    
    # Title Weight
    #####################################################
    title = (job.get("title") or "").lower()
    desc = (job.get("description") or "").lower()
    text = f"{title} {desc}"
    title_keywords = {"data analyst":30,"marketing analyst":45,"business analyst":30,"insights analyst":50,"reporting analyst":40,
        "bi analyst":30,"business intelligence":30, "sem":50,"ppc":50, "performance marketing":40}

    for keyword, pts in title_keywords.items():
        if keyword in title:
            score += pts

    #####################################################
    # Skills
    #####################################################
    text = f"{title} {desc}"
    skill_weights = {"Google Ads": 8,"Search ads":8,"sql": 8,"excel": 6,"tableau": 6,"statistics": 6,"dashboard": 5,
    "reporting": 5,"visualization": 5,"data analytics": 7,"data analysis": 7,"google analytics": 7,"paid search":8}
    
    for skill, pts in skill_weights.items():
        if skill in text:
          score += pts
    return score
    
   
def parse_jobs(raw: list[dict]) -> list[dict]:
    parsed = []
    for job in raw:
        try:
            dt     = datetime.fromisoformat(
                job.get("created", "").replace("Z", "+00:00"))
            posted = dt.strftime("%Y-%m-%d %H:%M UTC")
        except Exception:
            posted = "Unknown"

        area = (job.get("location") or {}).get("area", [])
        location = ", ".join(area) if area else "Unknown"
      
        sal_min = job.get("salary_min")
        sal_max = job.get("salary_max")

        if sal_min is not None and sal_max is not None:
             salary = f"${sal_min:,.0f} - ${sal_max:,.0f}"

        elif sal_min is not None:
             salary = f"${sal_min:,.0f}+"

        elif sal_max is not None:
             salary = f"Up to ${sal_max:,.0f}"

        else:
             salary = "Not listed"
              

        desc    = (job.get("description") or "")[:3000]
        title   = job.get("title") or ""
        company = (job.get("company") or {}).get("display_name") or ""
        matched = [s for s in SKILLS if s.lower() in (desc + title).lower()]
        remote = is_remote_job(title, desc)

        parsed.append({
            "Title":              title,
            "Company":            company,
            "Location":           location,
            "Remote":             "Yes" if remote else "No",
            "Salary":             salary,
            "Posted":             posted,
            "Skills Match":       ", ".join(matched),
            "Match Score":        score_job(job),
            "Key Qualifications": desc[:500],
            "Apply Link":         job.get("redirect_url") or "",
        })

    parsed.sort(key=lambda x: x["Match Score"], reverse=True)
    return parsed


# ── Excel Builder ──────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
ALT_FILL    = PatternFill("solid", start_color="D6E4F0")
HIGH_FILL   = PatternFill("solid", start_color="E2EFDA")

COLUMNS = [
    ("Title",                          35),
    ("Company",                        25),
    ("Location",                       20),
    ("Remote",                          9),
    ("Salary",                         22),
    ("Posted",                         20),
    ("Skills Match",                   30),
    ("Match Score",                    14),
    ("Key Qualifications",             45),
    ("Apply Link",                     15),
]

SCORE_COL = "Match Score"


def build_excel(jobs: list[dict], filepath: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title   = "Jobs"
    run_date   = datetime.now().strftime("%B %d, %Y")
    last_col   = get_column_letter(len(COLUMNS))
    thin = Border(
        left=Side(style="thin",   color="BFBFBF"),
        right=Side(style="thin",  color="BFBFBF"),
        top=Side(style="thin",    color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )

    # Title row
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = f"Paid Search Job Search  —  {run_date}  |  Past 24 Hours"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Subtitle row
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = (f"Skills: {', '.join(SKILLS)}   |   "
                f"Total roles found: {len(jobs)}")
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="595959")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    ws.append([])  # spacer

    # Header row
    headers = [c[0] for c in COLUMNS]
    ws.append(headers)
    for col_idx, (_, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = thin
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[4].height = 22

    # Data rows
    for row_num, job in enumerate(jobs, start=5):
        ws.append([job.get(h if h != SCORE_COL else "Match Score", "")
                   for h in headers])

        is_high = job.get("Match Score", 0) >= 55
        fill    = HIGH_FILL if is_high else (
                  ALT_FILL  if row_num % 2 == 0 else PatternFill())

        for col_idx in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.border = thin
            cell.font   = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center",
                                       wrap_text=(col_idx in (1, 9)))
            if fill:
                cell.fill = fill

        # Apply link hyperlink
        apply_col = headers.index("Apply Link") + 1
        url = job.get("Apply Link", "")
        if url:
            c = ws.cell(row=row_num, column=apply_col)
            c.hyperlink = url
            c.font  = Font(name="Arial", size=10, color="0563C1", underline="single")
            c.value = "Apply ->"

        # Center align score and remote
        ws.cell(row=row_num, column=headers.index(SCORE_COL)+1).alignment = \
            Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_num, column=headers.index("Remote")+1).alignment = \
            Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row_num].height = 32

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{last_col}{ws.max_row}"
    wb.save(filepath)
    print(f"  Excel saved: {filepath.name}")


# ── Email ──────────────────────────────────────────────────────────────────────

def send_email(excel_path: Path, job_count: int,
               gmail_address: str, app_password: str) -> None:
    today   = datetime.now().strftime("%B %d, %Y")
    subject = f"Paid Search Jobs — {job_count} Roles Found — {today}"
    body    = f"""Good morning!

Here are the Paid Search roles posted in the last 24 hours.

  Total roles found:      {job_count}
  Skills matched against: {', '.join(SKILLS)}

Open the attached Excel to browse all roles.
Green rows = strongest skill matches.
Click "Apply ->" in the last column to go directly to the job.

Good luck today!
"""
    msg = MIMEMultipart()
    msg["From"]    = gmail_address
    msg["To"]      = gmail_address
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))

    with open(excel_path, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition",
                    f"attachment; filename={excel_path.name}")
    msg.attach(part)

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(gmail_address, app_password)
        server.sendmail(gmail_address, gmail_address, msg.as_string())
    print(f"  Email sent to {gmail_address}")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    adzuna_app_id  = os.environ["ADZUNA_APP_ID"]
    adzuna_app_key = os.environ["ADZUNA_APP_KEY"]
    gmail_address  = os.environ["GMAIL_ADDRESS"]
    app_password   = os.environ["GMAIL_APP_PASSWORD"]

    today     = datetime.now().strftime("%Y-%m-%d")
    out_excel = OUTPUT_DIR / f"PaidSearch_Jobs_{today}.xlsx"

    print("=== Paid Search Job Search ===")
    print(f"Date: {today}\n")

    print("[1/3] Fetching jobs from Adzuna...")
    raw_jobs = fetch_jobs(adzuna_app_id, adzuna_app_key)

    print("\n[2/3] Parsing and scoring jobs...")
    jobs = parse_jobs(raw_jobs)
    print(f"  Jobs parsed: {len(jobs)}")

    print("\n[3/3] Building Excel and sending email...")
    build_excel(jobs, out_excel)
    send_email(out_excel, len(jobs), gmail_address, app_password)

    print(f"\nDone! {len(jobs)} jobs emailed to {gmail_address}")


if __name__ == "__main__":
    main()
